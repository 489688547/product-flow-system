#!/usr/bin/env python3
"""Merge split Kuaimai order-detail exports into a compact import CSV.

Only business aggregates are written. Customer/order-level fields are never
copied to the output.
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import date, datetime
from itertools import chain
from pathlib import Path

from openpyxl import load_workbook


BARCODE = re.compile(r"^69\d{10,12}$")
HEADER_TRIM = re.compile(r"[\s（(].*$")
SYNONYMS = {
    "code": ["规格商家编码", "系统商品编码", "商品编码", "69码", "条码", "条形码", "商品条码", "barcode"],
    "fallback_code": ["主商家编码"],
    "title": ["系统商品标题", "商品标题", "商品名称", "标题"],
    "platform": ["店铺平台", "所属平台", "平台", "渠道"],
    "create_time": ["订单创建时间", "创建时间", "下单时间", "交易创建时间"],
    "pay_time": ["付款时间", "支付时间"],
    "consign_time": ["发货时间"],
    "finish_time": ["完成时间", "成交时间"],
    "qty": ["净销量", "销量", "数量"],
    "gross_qty": ["销售数量"],
    "return_qty": ["退货数量"],
    "sales": ["订单买家已付金额", "商品买家已付金额", "买家已付金额", "已付金额", "支付金额", "销售额"],
    "net_sales": ["净销售额", "实收金额", "净销售"],
    "gross_sales": ["销售金额"],
    "gross_profit": ["净毛利"],
    "refund": ["退款金额", "退款"],
    "cost": ["销售成本", "成本"],
    "return_cost": ["退货成本"],
    "pre_ship_rate": ["发货前退款率"],
    "post_ship_rate": ["发货后退款率"],
}


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def header_text(value: object) -> str:
    return HEADER_TRIM.sub("", text(value))


def find_column(headers: list[object], names: list[str]) -> int | None:
    normalized = [header_text(value) for value in headers]
    for name in names:
        if name in normalized:
            return normalized.index(name)
    for name in names:
        for index, value in enumerate(normalized):
            if value and value.startswith(name):
                return index
    return None


def detect_mapping(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    for header_index, row in enumerate(rows[:8]):
        code_at = find_column(list(row), SYNONYMS["code"] + SYNONYMS["fallback_code"])
        qty_at = find_column(list(row), SYNONYMS["qty"] + SYNONYMS["gross_qty"])
        if code_at is None or qty_at is None:
            continue
        mapping = {}
        for key, names in SYNONYMS.items():
            at = find_column(list(row), names)
            if at is not None:
                mapping[key] = at
        if "net_sales" not in mapping and "gross_sales" not in mapping:
            raise ValueError("缺少净销售额/销售金额列")
        return header_index, mapping
    raise ValueError("前 8 行没有找到快麦销售明细表头")


def number(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    raw = text(value).replace(",", "").replace("¥", "").replace("￥", "").replace("%", "")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def rate(value: object) -> float:
    raw = text(value)
    parsed = number(value)
    return parsed / 100 if "%" in raw or parsed > 1.5 else parsed


def day(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    match = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text(value))
    if not match:
        return ""
    year, month, item_day = (int(part) for part in match.groups())
    if year < 2015 or year > 2100:
        return ""
    try:
        return date(year, month, item_day).isoformat()
    except ValueError:
        return ""


def rounded(value: float) -> float:
    return round(value + 1e-12, 2)


def aggregate(paths: list[Path]) -> tuple[dict[tuple[str, str, str], dict[str, object]], dict[str, str], int, int]:
    buckets: dict[tuple[str, str, str], dict[str, object]] = {}
    titles: dict[str, str] = {}
    source_rows = 0
    skipped = 0
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        head = []
        for _ in range(8):
            try:
                head.append(next(iterator))
            except StopIteration:
                break
        header_index, mapping = detect_mapping(head)
        rows = chain(head[header_index + 1 :], iterator)
        derives_net = "net_sales" not in mapping and "gross_sales" in mapping

        def cell(row: tuple[object, ...], key: str) -> object:
            index = mapping.get(key)
            return "" if index is None or index >= len(row) else row[index]

        for row in rows:
            source_rows += 1
            primary = text(cell(row, "code"))
            fallback = text(cell(row, "fallback_code"))
            code = primary if BARCODE.fullmatch(primary) else fallback
            created = (
                day(cell(row, "create_time"))
                or day(cell(row, "pay_time"))
                or day(cell(row, "consign_time"))
                or day(cell(row, "finish_time"))
            )
            if not BARCODE.fullmatch(code) or not created:
                skipped += 1
                continue
            platform = text(cell(row, "platform")) or "未知平台"
            key = (code, created, platform)
            bucket = buckets.setdefault(
                key,
                {
                    "qty": 0.0,
                    "sales": 0.0,
                    "net_sales": 0.0,
                    "gross_profit": 0.0,
                    "refund": 0.0,
                    "cost": 0.0,
                    "pre_ship_refund": 0.0,
                    "post_ship_refund": 0.0,
                },
            )
            sales = number(cell(row, "sales"))
            refund = number(cell(row, "refund"))
            qty = number(cell(row, "qty")) if "qty" in mapping else number(cell(row, "gross_qty")) - number(cell(row, "return_qty"))
            net_sales = number(cell(row, "gross_sales")) - refund if derives_net else number(cell(row, "net_sales"))
            cost = number(cell(row, "cost")) - number(cell(row, "return_cost")) if derives_net else number(cell(row, "cost"))
            gross_profit = net_sales - cost if "gross_profit" not in mapping and derives_net else number(cell(row, "gross_profit"))
            bucket["qty"] += qty
            bucket["sales"] += sales
            bucket["net_sales"] += net_sales
            bucket["gross_profit"] += gross_profit
            bucket["refund"] += refund
            bucket["cost"] += cost
            bucket["pre_ship_refund"] += rate(cell(row, "pre_ship_rate")) * sales
            bucket["post_ship_refund"] += rate(cell(row, "post_ship_rate")) * sales
            item_title = text(cell(row, "title"))
            if item_title and code not in titles:
                titles[code] = item_title
        workbook.close()
    return buckets, titles, source_rows, skipped


def write_csv(output: Path, buckets: dict[tuple[str, str, str], dict[str, object]], titles: dict[str, str]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([
            "规格商家编码",
            "商品标题",
            "所属平台",
            "下单时间",
            "净销量",
            "商品买家已付金额",
            "净销售额",
            "净毛利",
            "退款金额",
            "销售成本",
            "发货前退款率",
            "发货后退款率",
        ])
        for (code, created, platform), values in sorted(buckets.items(), key=lambda item: (item[0][1], item[0][0], item[0][2])):
            sales = float(values["sales"])
            writer.writerow([
                code,
                titles.get(code, ""),
                platform,
                created,
                round(float(values["qty"])),
                rounded(sales),
                rounded(float(values["net_sales"])),
                rounded(float(values["gross_profit"])),
                rounded(float(values["refund"])),
                rounded(float(values["cost"])),
                round(float(values["pre_ship_refund"]) / sales, 6) if sales else 0,
                round(float(values["post_ship_refund"]) / sales, 6) if sales else 0,
            ])


def split_compact_csv(source: Path, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    handles: dict[str, object] = {}
    writers: dict[str, csv.DictWriter] = {}
    outputs: list[Path] = []
    try:
        with source.open("r", encoding="utf-8", newline="") as input_handle:
            reader = csv.DictReader(input_handle)
            if not reader.fieldnames or "下单时间" not in reader.fieldnames:
                raise ValueError("汇总 CSV 缺少下单时间列")
            for row in reader:
                month = text(row.get("下单时间"))[:7]
                if not re.fullmatch(r"\d{4}-\d{2}", month):
                    continue
                if month not in writers:
                    output = output_dir / f"{month}.csv"
                    handle = output.open("w", encoding="utf-8", newline="")
                    writer = csv.DictWriter(handle, fieldnames=reader.fieldnames)
                    writer.writeheader()
                    handles[month] = handle
                    writers[month] = writer
                    outputs.append(output)
                writers[month].writerow(row)
    finally:
        for handle in handles.values():
            handle.close()
    return sorted(outputs)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--split-compact", type=Path)
    parser.add_argument("inputs", nargs="*", type=Path)
    args = parser.parse_args()
    if args.split_compact:
        outputs = split_compact_csv(args.split_compact, args.output)
        print({"outputs": [str(path) for path in outputs]})
        return
    if not args.inputs:
        parser.error("至少需要一个 Excel 输入文件")
    buckets, titles, source_rows, skipped = aggregate(args.inputs)
    write_csv(args.output, buckets, titles)
    months = sorted({created[:7] for _, created, _ in buckets})
    print({"source_rows": source_rows, "skipped": skipped, "aggregate_rows": len(buckets), "months": months, "output": str(args.output)})


if __name__ == "__main__":
    main()
